#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
core.excel_table
Создание и обновление Excel-таблиц по нормализованным данным фида.

Главный принцип обновления:
- XML — источник истины для динамических полей;
- ручные поля сохраняются строго по ID квартиры, а не по номеру строки;
- итоговая таблица пересобирается заново с первой строки, поэтому пустые строки и сдвиги не сохраняются.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Не установлен openpyxl. Установите зависимость командой: python3 -m pip install openpyxl"
    ) from exc

from .feed_parser import COLUMNS_BASE, build_columns, normalize_cell_value

EXCEL_ENGINE_VERSION = "0.5-modular-rebuild-by-id"

# Эти поля всегда обновляются из XML для существующих строк.
DYNAMIC_UPDATE_FIELDS = {
    "Телефон",
    "Отделка",
    "Цена",
    "Цена со скидкой",
    "Описание",
    "Планировка",
}

# Эти поля при обновлении сохраняются из старой таблицы по ID, если в старой таблице есть значение.
MANUAL_OR_STATIC_FIELDS = {
    "Название ЖК",
    "Корпус",
    "Номер квартиры",
    "Наименование застройщика",
    "Адрес",
    "Количество комнат",
    "Этаж",
    "Этажность дома",
    "Общая площадь",
    "Проектная декларация",
    "Главное изображение",
}


def normalize_id(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text


def is_blank(value: Any) -> bool:
    return value is None or str(value).strip() == ""


def image_columns(columns: list[str]) -> list[str]:
    def image_num(name: str) -> int:
        try:
            return int(name.replace("Изображение", "").strip())
        except Exception:
            return 10**9

    return sorted([c for c in columns if c.startswith("Изображение ")], key=image_num)


def build_output_columns(old_columns: list[str], feed_columns: list[str]) -> list[str]:
    """Собирает итоговый набор колонок без дубликатов.

    Базовые колонки всегда идут первыми. Потом — изображения в нормальном порядке.
    Дополнительные пользовательские колонки из старой таблицы сохраняются между базой и изображениями.
    """
    result: list[str] = []

    for col in COLUMNS_BASE:
        if col not in result:
            result.append(col)

    old_image_cols = set(image_columns(old_columns))
    feed_image_cols = set(image_columns(feed_columns))
    all_image_cols = sorted(old_image_cols | feed_image_cols, key=lambda c: int(c.replace("Изображение", "").strip()))

    for col in old_columns:
        if col not in result and not col.startswith("Изображение "):
            result.append(col)

    for col in feed_columns:
        if col not in result and not col.startswith("Изображение "):
            result.append(col)

    for col in all_image_cols:
        if col not in result:
            result.append(col)

    return result


def style_header_cell(cell) -> None:
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1F4E78")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin = Side(style="thin", color="D9E2F3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def style_body_cell(cell) -> None:
    cell.alignment = Alignment(vertical="top", wrap_text=True)
    thin = Side(style="thin", color="D9E2F3")
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)


def set_column_width(ws, col_idx: int, col_name: str) -> None:
    widths = {
        "ID квартиры": 14,
        "Название ЖК": 18,
        "Корпус": 22,
        "Номер квартиры": 16,
        "Наименование застройщика": 26,
        "Телефон": 22,
        "Адрес": 34,
        "Количество комнат": 16,
        "Отделка": 22,
        "Цена": 15,
        "Цена со скидкой": 18,
        "Этаж": 10,
        "Этажность дома": 16,
        "Общая площадь": 16,
        "Описание": 45,
        "Проектная декларация": 28,
        "Главное изображение": 35,
        "Планировка": 35,
    }
    width = widths.get(col_name, 35 if col_name.startswith("Изображение ") else 18)
    ws.column_dimensions[get_column_letter(col_idx)].width = width


def apply_formatting(ws, columns: list[str]) -> None:
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_idx, col_name in enumerate(columns, start=1):
        style_header_cell(ws.cell(row=1, column=col_idx))
        set_column_width(ws, col_idx, col_name)

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            style_body_cell(cell)


def write_rows_to_xlsx(rows: list[dict[str, Any]], columns: list[str], output: str | Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Квартиры"

    for col_idx, col_name in enumerate(columns, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    for row_idx, row_data in enumerate(rows, start=2):
        for col_idx, col_name in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=normalize_cell_value(row_data.get(col_name, "")))

    apply_formatting(ws, columns)
    wb.save(output)


def create_excel_from_feed_rows(rows: list[dict[str, Any]], feed_columns: list[str], output: str | Path) -> dict[str, Any]:
    # Важно: parse_universal() уже вызвал build_columns() и разложил изображения по колонкам.
    # Повторно вызывать build_columns() нельзя, иначе изображения пропадут.
    columns = list(feed_columns)
    write_rows_to_xlsx(rows, columns, output)
    return {"rows": len(rows), "columns": len(columns), "output": str(output)}


def header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for cell in ws[1]:
        if cell.value is not None and str(cell.value).strip():
            headers[str(cell.value).strip()] = cell.column
    return headers


def row_values_by_header(ws, headers: dict[str, int], row_num: int) -> dict[str, Any]:
    return {name: ws.cell(row=row_num, column=col_idx).value for name, col_idx in headers.items()}


def read_existing_rows_by_id(excel_path: str | Path) -> tuple[dict[str, dict[str, Any]], list[str], int]:
    wb = load_workbook(excel_path)
    ws = wb.active
    headers = header_map(ws)

    if "ID квартиры" not in headers:
        raise ValueError('В Excel не найден обязательный столбец "ID квартиры"')

    existing: dict[str, dict[str, Any]] = {}
    filled_rows = 0
    id_col = headers["ID квартиры"]

    for row_num in range(2, ws.max_row + 1):
        apt_id = normalize_id(ws.cell(row=row_num, column=id_col).value)
        if not apt_id:
            continue
        filled_rows += 1
        existing[apt_id] = row_values_by_header(ws, headers, row_num)

    old_columns = list(headers.keys())
    return existing, old_columns, filled_rows


def make_updated_row(feed_row: dict[str, Any], old_row: dict[str, Any] | None, output_columns: list[str]) -> tuple[dict[str, Any], int]:
    """Создает строку для новой таблицы и возвращает количество измененных ячеек."""
    result: dict[str, Any] = {}
    changed_cells = 0

    for col in output_columns:
        feed_value = feed_row.get(col, "")
        old_value = old_row.get(col, "") if old_row else ""

        if col == "ID квартиры":
            value = feed_row.get(col, old_value)
        elif col.startswith("Изображение "):
            value = feed_value
        elif col in DYNAMIC_UPDATE_FIELDS:
            value = feed_value
        elif col in MANUAL_OR_STATIC_FIELDS:
            # Для существующих строк сохраняем ручную/статичную информацию, если она есть.
            # Для новых строк берем данные из XML.
            value = old_value if old_row and not is_blank(old_value) else feed_value
        else:
            # Пользовательские дополнительные колонки, если появятся в таблице, сохраняем по ID.
            value = old_value if old_row else feed_value

        result[col] = normalize_cell_value(value)

        if old_row is not None:
            old_norm = "" if old_value is None else str(old_value).strip()
            new_norm = "" if value is None else str(value).strip()
            if old_norm != new_norm and (col in DYNAMIC_UPDATE_FIELDS or col.startswith("Изображение ")):
                changed_cells += 1

    return result, changed_cells


def update_excel_from_feed_rows(rows: list[dict[str, Any]], feed_columns: list[str], excel_path: str | Path, output_path: str | Path) -> dict[str, Any]:
    existing_by_id, old_columns, excel_rows_before = read_existing_rows_by_id(excel_path)

    feed_by_id: dict[str, dict[str, Any]] = {}
    feed_order: list[str] = []
    for row in rows:
        apt_id = normalize_id(row.get("ID квартиры"))
        if not apt_id:
            continue
        if apt_id not in feed_by_id:
            feed_order.append(apt_id)
        feed_by_id[apt_id] = row

    output_columns = build_output_columns(old_columns, feed_columns)

    output_rows: list[dict[str, Any]] = []
    added = 0
    updated_apartments = 0
    updated_cells = 0
    prices_changed = 0
    added_ids: list[str] = []
    price_changes: list[dict[str, str]] = []

    for apt_id in feed_order:
        feed_row = feed_by_id[apt_id]
        old_row = existing_by_id.get(apt_id)
        if old_row is None:
            added += 1
            added_ids.append(apt_id)
        new_row, changed_cells = make_updated_row(feed_row, old_row, output_columns)
        if changed_cells > 0:
            updated_apartments += 1
            updated_cells += changed_cells
        if old_row is not None:
            old_price = "" if old_row.get("Цена") is None else str(old_row.get("Цена")).strip()
            new_price = "" if feed_row.get("Цена") is None else str(feed_row.get("Цена")).strip()
            if old_price != new_price:
                prices_changed += 1
                price_changes.append({
                    "ID квартиры": apt_id,
                    "old_price": old_price,
                    "new_price": new_price,
                })
        output_rows.append(new_row)

    deleted_ids = sorted(set(existing_by_id) - set(feed_by_id))
    deleted = len(deleted_ids)

    write_rows_to_xlsx(output_rows, output_columns, output_path)

    return {
        "engine_version": EXCEL_ENGINE_VERSION,
        "excel_rows_before": excel_rows_before,
        "feed_rows": len(feed_order),
        "deleted": deleted,
        "added": added,
        "updated_apartments": updated_apartments,
        "updated_cells": updated_cells,
        "prices_changed": prices_changed,
        "added_ids": added_ids,
        "deleted_ids": deleted_ids,
        "price_changes": price_changes,
        "excel_rows_after": len(output_rows),
        "output": str(output_path),
    }
