#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Универсальное обновление Excel-таблицы по XML-фиду недвижимости.

Поддерживаемые форматы XML:
1) Яндекс.Недвижимость: <realty-feed> ... <offer internal-id="...">
2) Домклик / Kvartirogramma: <complexes> ... <complex> ... <buildings> ... <flat>

Что делает скрипт:
- читает XML из файла или по ссылке;
- автоматически определяет формат фида;
- сопоставляет строки Excel и квартиры XML по полю "ID квартиры";
- удаляет строки, ID которых больше нет в XML;
- добавляет новые квартиры из XML;
- обновляет только динамические поля из XML;
- ручные поля существующих строк не трогает.

Запуск из VS Code / терминала:
    python3 universal_update_excel_from_feed.py --source feed.xml --excel lots.xlsx --output lots_updated.xlsx

Запуск по ссылке:
    python3 universal_update_excel_from_feed.py --source "https://example.com/feed.xml" --excel lots.xlsx --output lots_updated.xlsx

Требования:
    python3 -m pip install openpyxl requests
"""

from __future__ import annotations

import argparse
import copy
import sys
import urllib.request
import xml.etree.ElementTree as ET
from pathlib import Path
from typing import Iterable, Optional

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError as exc:
    raise SystemExit(
        "Не установлен openpyxl. Установите зависимость командой: python3 -m pip install openpyxl"
    ) from exc


COLUMNS_BASE = [
    "ID квартиры",
    "Название ЖК",
    "Корпус",
    "Номер квартиры",
    "Наименование застройщика",
    "Телефон",
    "Адрес",
    "Количество комнат",
    "Отделка",
    "Цена",
    "Цена со скидкой",
    "Этаж",
    "Этажность дома",
    "Общая площадь",
    "Описание",
    "Проектная декларация",
    "Главное изображение",
    "Планировка",
]

# Эти поля у существующих строк всегда берутся из XML.
# Важно: "Телефон" тоже обновляется из XML, как вы уточнили.
DYNAMIC_UPDATE_FIELDS = {
    "Телефон",
    "Отделка",
    "Цена",
    "Цена со скидкой",
    "Описание",
    "Планировка",
}

# Эти поля у существующих строк НЕ перезаписываются, чтобы не стереть ручную доработку.
# Для новых квартир они заполняются из XML, если данные есть.
MANUAL_OR_STATIC_FIELDS = {
    "Название ЖК",
    "Корпус",
    "Номер квартиры",
    "Наименование застройщика",
    "Адрес",
    "Количество комнат",
    "Этаж",
    "Этажность дома",
    "Общая площадь",
    "Проектная декларация",
    "Главное изображение",
}

YANDEX_FORMAT = "yandex"
DOMCLICK_FORMAT = "domclick"
SCRIPT_VERSION = "0.4-rebuild-by-id"


def strip_ns(tag: str) -> str:
    return tag.split("}", 1)[-1] if "}" in tag else tag


def get_namespace(root: ET.Element) -> str:
    if root.tag.startswith("{"):
        return root.tag.split("}", 1)[0].strip("{")
    return ""


def q(ns: str, tag: str) -> str:
    return f"{{{ns}}}{tag}" if ns else tag


def child_text(parent: Optional[ET.Element], tag: str, default: str = "") -> str:
    if parent is None:
        return default
    child = parent.find(tag)
    if child is None or child.text is None:
        return default
    return child.text.strip()


def text(parent: ET.Element, ns: str, path: str, default: str = "") -> str:
    current: Optional[ET.Element] = parent
    for part in path.split("/"):
        if current is None:
            return default
        current = current.find(q(ns, part))
    if current is None or current.text is None:
        return default
    return current.text.strip()


def first_existing_text(parent: ET.Element, ns: str, paths: Iterable[str], default: str = "") -> str:
    for path in paths:
        value = text(parent, ns, path, "")
        if value != "":
            return value
    return default


def normalize_id(value: object) -> str:
    if value is None:
        return ""
    # Excel иногда хранит целые ID как число/float. Приводим к стабильной строке.
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def normalize_cell_value(value: object) -> object:
    if not isinstance(value, str):
        return value
    value = value.strip()
    if value == "":
        return ""
    try:
        if value.replace(".", "", 1).isdigit():
            return float(value) if "." in value else int(value)
    except Exception:
        pass
    return value


def load_xml(source: str) -> bytes:
    if source.startswith(("http://", "https://")):
        try:
            import requests  # type: ignore
            response = requests.get(source, timeout=60)
            response.raise_for_status()
            return response.content
        except ImportError:
            with urllib.request.urlopen(source, timeout=60) as response:
                return response.read()
    return Path(source).read_bytes()


def detect_feed_format(root: ET.Element) -> str:
    root_name = strip_ns(root.tag).lower()
    if root_name == "complexes" and root.findall(".//complex"):
        return DOMCLICK_FORMAT

    ns = get_namespace(root)
    if root.findall(f".//{q(ns, 'offer')}"):
        return YANDEX_FORMAT

    raise ValueError(
        "Не удалось определить формат XML. Поддерживаются Яндекс.Недвижимость и Домклик/Kvartirogramma."
    )


def build_columns(rows: list[dict[str, object]]) -> list[str]:
    max_images = 0
    for row in rows:
        regular_images = row.get("__regular_images", [])
        if isinstance(regular_images, list):
            max_images = max(max_images, len(regular_images))

    columns = COLUMNS_BASE + [f"Изображение {i}" for i in range(1, max_images + 1)]

    for row in rows:
        regular_images = row.pop("__regular_images", [])
        if not isinstance(regular_images, list):
            regular_images = []
        for i in range(1, max_images + 1):
            row[f"Изображение {i}"] = regular_images[i - 1] if i <= len(regular_images) else ""

    return columns


def parse_yandex(root: ET.Element) -> tuple[list[dict[str, object]], list[str]]:
    ns = get_namespace(root)
    offers = root.findall(f".//{q(ns, 'offer')}")
    rows: list[dict[str, object]] = []

    for offer in offers:
        plan_image = ""
        regular_images: list[str] = []
        first_offer_image = ""

        for image in offer.findall(q(ns, "image")):
            url = (image.text or "").strip()
            if not url:
                continue
            if not first_offer_image:
                first_offer_image = url

            image_tag = image.attrib.get("tag", "")
            if image_tag == "plan":
                if not plan_image:
                    plan_image = url
            elif image_tag == "floor-plan":
                continue
            else:
                regular_images.append(url)

        # Если отдельного тега планировки нет, планировкой считаем первое изображение квартиры.
        if not plan_image:
            plan_image = first_offer_image

        rows.append({
            "ID квартиры": offer.attrib.get("internal-id", ""),
            "Название ЖК": text(offer, ns, "building-name"),
            "Корпус": text(offer, ns, "building-section"),
            "Номер квартиры": text(offer, ns, "apartments"),
            "Наименование застройщика": first_existing_text(
                offer, ns, ["sales-agent/organization", "sales-agent/name"]
            ),
            "Телефон": text(offer, ns, "sales-agent/phone"),
            "Адрес": text(offer, ns, "location/address"),
            "Количество комнат": text(offer, ns, "rooms"),
            "Отделка": text(offer, ns, "renovation"),
            "Цена": text(offer, ns, "price/value"),
            "Цена со скидкой": first_existing_text(
                offer,
                ns,
                [
                    "discount-price/value",
                    "price-with-discount/value",
                    "sale-price/value",
                    "price-discount/value",
                ],
            ),
            "Этаж": text(offer, ns, "floor"),
            "Этажность дома": text(offer, ns, "floors-total"),
            "Общая площадь": text(offer, ns, "area/value"),
            "Описание": text(offer, ns, "description"),
            "Проектная декларация": "",
            "Главное изображение": "",
            "Планировка": plan_image,
            "__regular_images": regular_images,
        })

    return rows, build_columns(rows)


def domclick_complex_description(complex_el: ET.Element) -> str:
    pieces: list[str] = []
    direct_description = child_text(complex_el, "description")
    if direct_description:
        pieces.append(direct_description)

    for desc in complex_el.findall("description_secondary"):
        title = child_text(desc, "title")
        body = child_text(desc, "text")
        if title and body:
            pieces.append(f"{title}. {body}")
        elif body:
            pieces.append(body)

    return "\n\n".join(pieces).strip()


def domclick_phones(complex_el: ET.Element) -> str:
    candidates = [
        child_text(complex_el, "sales_info/sales_phone"),
        child_text(complex_el, "sales_info/responsible_officer_phone"),
        child_text(complex_el, "developer/phone"),
    ]

    # На случай нестандартной структуры: собираем все теги, где в названии есть phone.
    for elem in complex_el.iter():
        tag = strip_ns(elem.tag).lower()
        if "phone" in tag and elem.text and elem.text.strip():
            candidates.append(elem.text.strip())

    unique: list[str] = []
    for phone in candidates:
        phone = (phone or "").strip()
        if phone and phone not in unique:
            unique.append(phone)
    return "; ".join(unique)


def parse_domclick(root: ET.Element) -> tuple[list[dict[str, object]], list[str]]:
    rows: list[dict[str, object]] = []

    for complex_el in root.findall("complex"):
        complex_name = child_text(complex_el, "name")
        complex_address = child_text(complex_el, "address")
        complex_images = [
            (img.text or "").strip()
            for img in complex_el.findall("images/image")
            if img.text and img.text.strip()
        ]
        complex_description = domclick_complex_description(complex_el)
        phone = domclick_phones(complex_el)
        developer_name = child_text(complex_el, "developer/name")

        for building in complex_el.findall("buildings/building"):
            building_name = child_text(building, "name")
            building_address = child_text(building, "address") or complex_address
            floors_total = child_text(building, "floors")

            for flat in building.findall("flats/flat"):
                flat_plans = [
                    (plan.text or "").strip()
                    for plan in flat.findall("plans/plan")
                    if plan.text and plan.text.strip()
                ]

                regular_images = flat_plans + complex_images
                plan_image = flat_plans[0] if flat_plans else ""

                rows.append({
                    "ID квартиры": child_text(flat, "flat_id"),
                    "Название ЖК": complex_name,
                    "Корпус": building_name,
                    "Номер квартиры": child_text(flat, "apartment"),
                    "Наименование застройщика": developer_name,
                    "Телефон": phone,
                    "Адрес": building_address,
                    "Количество комнат": child_text(flat, "room"),
                    "Отделка": child_text(flat, "renovation"),
                    "Цена": child_text(flat, "price"),
                    "Цена со скидкой": "",
                    "Этаж": child_text(flat, "floor"),
                    "Этажность дома": floors_total,
                    "Общая площадь": child_text(flat, "area"),
                    "Описание": complex_description,
                    "Проектная декларация": "",
                    "Главное изображение": "",
                    "Планировка": plan_image,
                    "__regular_images": regular_images,
                })

    return rows, build_columns(rows)


def parse_universal(xml_bytes: bytes) -> tuple[list[dict[str, object]], list[str], str]:
    root = ET.fromstring(xml_bytes)
    feed_format = detect_feed_format(root)

    if feed_format == YANDEX_FORMAT:
        rows, columns = parse_yandex(root)
    elif feed_format == DOMCLICK_FORMAT:
        rows, columns = parse_domclick(root)
    else:
        raise ValueError(f"Неподдерживаемый формат: {feed_format}")

    # Проверка дублей ID в XML: для обновления это критично.
    seen: set[str] = set()
    duplicates: list[str] = []
    for row in rows:
        row_id = normalize_id(row.get("ID квартиры"))
        if not row_id:
            continue
        if row_id in seen:
            duplicates.append(row_id)
        seen.add(row_id)
    if duplicates:
        sample = ", ".join(duplicates[:10])
        raise ValueError(f"В XML найдены повторяющиеся ID квартир: {sample}")

    return rows, columns, feed_format


def header_map(ws) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=1, column=col).value
        if value is not None:
            headers[str(value).strip()] = col
    return headers


def ensure_column(ws, headers: dict[str, int], col_name: str) -> int:
    if col_name in headers:
        return headers[col_name]
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col).value = col_name
    headers[col_name] = new_col
    style_header_cell(ws.cell(row=1, column=new_col))
    set_column_width(ws, new_col, col_name)
    return new_col


def image_columns_from_headers(headers: dict[str, int]) -> list[str]:
    result: list[tuple[int, str]] = []
    for name in headers:
        if name.startswith("Изображение "):
            try:
                number = int(name.replace("Изображение ", "").strip())
                result.append((number, name))
            except ValueError:
                pass
    return [name for _, name in sorted(result)]


def get_existing_row_map(ws, headers: dict[str, int]) -> dict[str, int]:
    if "ID квартиры" not in headers:
        raise ValueError('В Excel не найден обязательный столбец "ID квартиры"')

    id_col = headers["ID квартиры"]
    mapping: dict[str, int] = {}
    duplicates: list[str] = []

    for row_num in range(2, ws.max_row + 1):
        row_id = normalize_id(ws.cell(row=row_num, column=id_col).value)
        if not row_id:
            continue
        if row_id in mapping:
            duplicates.append(row_id)
        mapping[row_id] = row_num

    if duplicates:
        sample = ", ".join(duplicates[:10])
        raise ValueError(f"В Excel найдены повторяющиеся ID квартир: {sample}")

    return mapping


def is_row_fully_empty(ws, row_num: int) -> bool:
    """Проверяет, что строка полностью пустая по значениям ячеек.

    Форматирование не учитывается: если в строке нет ни одного значения,
    такую строку можно безопасно удалить.
    """
    for col in range(1, ws.max_column + 1):
        value = ws.cell(row=row_num, column=col).value
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return False
    return True


def delete_fully_empty_rows(ws) -> int:
    """Удаляет полностью пустые строки ниже заголовка и возвращает их количество."""
    deleted = 0
    for row_num in range(ws.max_row, 1, -1):
        if is_row_fully_empty(ws, row_num):
            ws.delete_rows(row_num, 1)
            deleted += 1
    return deleted


def delete_rows_without_id(ws, headers: dict[str, int]) -> int:
    """Удаляет строки таблицы, в которых нет ID квартиры.

    В реальной работе пользователь может не удалить строку Excel физически,
    а просто очистить ее содержимое. Из-за форматирования/служебных хвостов Excel
    такая строка не всегда считается «полностью пустой».

    Для нашей таблицы строка без ID квартиры не является квартирой, поэтому
    ниже заголовка такие строки безопасно удаляем.
    """
    id_col = headers.get("ID квартиры")
    if not id_col:
        return 0

    deleted = 0
    for row_num in range(ws.max_row, 1, -1):
        row_id = normalize_id(ws.cell(row=row_num, column=id_col).value)
        if row_id:
            continue
        ws.delete_rows(row_num, 1)
        deleted += 1
    return deleted


def count_rows_with_id(ws, headers: dict[str, int]) -> int:
    """Считает реальные строки квартир по заполненному столбцу ID квартиры."""
    id_col = headers.get("ID квартиры")
    if not id_col:
        return 0
    count = 0
    for row_num in range(2, ws.max_row + 1):
        if normalize_id(ws.cell(row=row_num, column=id_col).value):
            count += 1
    return count


def clear_row_values(ws, row_num: int, start_col: int, end_col: int) -> None:
    for col in range(start_col, end_col + 1):
        ws.cell(row=row_num, column=col).value = None


def copy_row_style(ws, source_row: int, target_row: int) -> None:
    if source_row < 1 or source_row > ws.max_row:
        return
    for col in range(1, ws.max_column + 1):
        source_cell = ws.cell(row=source_row, column=col)
        target_cell = ws.cell(row=target_row, column=col)
        if source_cell.has_style:
            target_cell._style = copy.copy(source_cell._style)
        if source_cell.number_format:
            target_cell.number_format = source_cell.number_format
        if source_cell.alignment:
            target_cell.alignment = copy.copy(source_cell.alignment)
        if source_cell.border:
            target_cell.border = copy.copy(source_cell.border)
        if source_cell.fill:
            target_cell.fill = copy.copy(source_cell.fill)
        if source_cell.font:
            target_cell.font = copy.copy(source_cell.font)


def write_full_row(ws, row_num: int, headers: dict[str, int], row_data: dict[str, object], all_columns: list[str]) -> None:
    for col_name in all_columns:
        col_idx = ensure_column(ws, headers, col_name)
        ws.cell(row=row_num, column=col_idx).value = normalize_cell_value(row_data.get(col_name, ""))


def update_existing_row(
    ws,
    row_num: int,
    headers: dict[str, int],
    row_data: dict[str, object],
    max_image_count: int,
) -> int:
    changes = 0

    # Динамические не-изображения.
    for col_name in sorted(DYNAMIC_UPDATE_FIELDS):
        col_idx = ensure_column(ws, headers, col_name)
        old_value = ws.cell(row=row_num, column=col_idx).value
        new_value = normalize_cell_value(row_data.get(col_name, ""))
        if normalize_compare_value(old_value) != normalize_compare_value(new_value):
            ws.cell(row=row_num, column=col_idx).value = new_value
            changes += 1

    # Изображения полностью перезаписываем из XML.
    for i in range(1, max_image_count + 1):
        col_name = f"Изображение {i}"
        col_idx = ensure_column(ws, headers, col_name)
        old_value = ws.cell(row=row_num, column=col_idx).value
        new_value = row_data.get(col_name, "")
        if normalize_compare_value(old_value) != normalize_compare_value(new_value):
            ws.cell(row=row_num, column=col_idx).value = new_value
            changes += 1

    # Если в старом Excel было больше колонок изображений, чем сейчас в XML,
    # очищаем лишние старые изображения, чтобы не оставались устаревшие ссылки.
    for image_col_name in image_columns_from_headers(headers):
        try:
            number = int(image_col_name.replace("Изображение ", "").strip())
        except ValueError:
            continue
        if number > max_image_count:
            col_idx = headers[image_col_name]
            if ws.cell(row=row_num, column=col_idx).value not in (None, ""):
                ws.cell(row=row_num, column=col_idx).value = None
                changes += 1

    return changes


def normalize_compare_value(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def style_header_cell(cell) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border


def style_body_row(ws, row_num: int) -> None:
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[row_num]:
        cell.alignment = Alignment(vertical="top", wrap_text=True)
        cell.border = border


def set_column_width(ws, col_idx: int, col_name: str) -> None:
    width_by_name = {
        "ID квартиры": 14,
        "Название ЖК": 20,
        "Корпус": 30,
        "Номер квартиры": 16,
        "Наименование застройщика": 26,
        "Телефон": 24,
        "Адрес": 38,
        "Количество комнат": 16,
        "Отделка": 22,
        "Цена": 14,
        "Цена со скидкой": 18,
        "Этаж": 10,
        "Этажность дома": 16,
        "Общая площадь": 16,
        "Описание": 50,
        "Проектная декларация": 28,
        "Главное изображение": 32,
        "Планировка": 42,
    }
    letter = get_column_letter(col_idx)
    ws.column_dimensions[letter].width = 42 if col_name.startswith("Изображение ") else width_by_name.get(col_name, 18)


def apply_basic_formatting(ws, headers: dict[str, int]) -> None:
    for col_idx in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col_idx)
        style_header_cell(cell)
        set_column_width(ws, col_idx, str(cell.value or ""))

    for row_num in range(2, ws.max_row + 1):
        style_body_row(ws, row_num)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_name in ("Цена", "Цена со скидкой"):
        if col_name in headers:
            col_idx = headers[col_name]
            for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in col_cells:
                    c.number_format = '#,##0'

    if "Общая площадь" in headers:
        col_idx = headers["Общая площадь"]
        for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in col_cells:
                c.number_format = '0.00'



def row_values_by_header(ws, headers: dict[str, int], row_num: int) -> dict[str, object]:
    result: dict[str, object] = {}
    for name, col_idx in headers.items():
        result[name] = ws.cell(row=row_num, column=col_idx).value
    return result


def count_rows_without_id_in_source(ws, headers: dict[str, int]) -> int:
    """Считает строки ниже заголовка, где нет ID квартиры.

    Такие строки не считаются квартирами и при обновлении не переносятся
    в новый файл. Это решает проблему «дыр» в таблице, когда пользователь
    очистил строки, но Excel продолжает хранить их как часть диапазона.
    """
    id_col = headers.get("ID квартиры")
    if not id_col:
        return 0
    count = 0
    for row_num in range(2, ws.max_row + 1):
        row_id = normalize_id(ws.cell(row=row_num, column=id_col).value)
        if not row_id:
            count += 1
    return count


def build_existing_data_by_id(ws, headers: dict[str, int]) -> dict[str, dict[str, object]]:
    """Читает старую Excel-таблицу и возвращает данные строк по ID квартиры."""
    if "ID квартиры" not in headers:
        raise ValueError('В Excel не найден обязательный столбец "ID квартиры"')

    id_col = headers["ID квартиры"]
    result: dict[str, dict[str, object]] = {}
    duplicates: list[str] = []

    for row_num in range(2, ws.max_row + 1):
        row_id = normalize_id(ws.cell(row=row_num, column=id_col).value)
        if not row_id:
            continue
        if row_id in result:
            duplicates.append(row_id)
            continue
        result[row_id] = row_values_by_header(ws, headers, row_num)

    if duplicates:
        sample = ", ".join(duplicates[:10])
        raise ValueError(f"В Excel найдены повторяющиеся ID квартир: {sample}")

    return result


def build_output_columns(old_headers: dict[str, int], feed_columns: list[str]) -> list[str]:
    """Формирует итоговый набор колонок.

    Базовые колонки идут в нашем фиксированном порядке, изображения — по числу
    изображений в актуальном XML. Дополнительные ручные колонки из старого Excel,
    если они были добавлены пользователем, сохраняются в конце.
    """
    max_image_count = max(
        [0]
        + [
            int(col.replace("Изображение ", ""))
            for col in feed_columns
            if col.startswith("Изображение ") and col.replace("Изображение ", "").isdigit()
        ]
    )
    base = COLUMNS_BASE + [f"Изображение {i}" for i in range(1, max_image_count + 1)]
    extra = [name for name in old_headers.keys() if name and name not in base]
    return base + extra


def make_output_row(
    row_id: str,
    feed_row: dict[str, object],
    old_row: Optional[dict[str, object]],
    output_columns: list[str],
) -> list[object]:
    """Собирает одну итоговую строку.

    Правила:
    - ID, динамические поля, телефон, планировка и изображения всегда берутся из XML;
    - ручные/статичные поля у существующих квартир сохраняются из старой таблицы;
    - у новых квартир статичные поля заполняются из XML, ручные остаются пустыми;
    - дополнительные пользовательские колонки сохраняются из старой таблицы.
    """
    values: list[object] = []
    old_row = old_row or {}

    for col_name in output_columns:
        if col_name == "ID квартиры":
            value = row_id
        elif col_name in DYNAMIC_UPDATE_FIELDS or col_name.startswith("Изображение "):
            value = feed_row.get(col_name, "")
        elif col_name in MANUAL_OR_STATIC_FIELDS:
            if old_row:
                value = old_row.get(col_name, "")
            else:
                # Для новых квартир ручные поля остаются пустыми, а статичные можно заполнить из XML.
                if col_name in {"Проектная декларация", "Главное изображение"}:
                    value = ""
                else:
                    value = feed_row.get(col_name, "")
        elif col_name in COLUMNS_BASE:
            value = feed_row.get(col_name, "")
        else:
            # Любые дополнительные пользовательские колонки сохраняем.
            value = old_row.get(col_name, "") if old_row else ""

        values.append(normalize_cell_value(value))

    return values


def apply_formatting_to_new_workbook(ws, output_columns: list[str]) -> dict[str, int]:
    headers: dict[str, int] = {}
    for col_idx, col_name in enumerate(output_columns, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = col_name
        style_header_cell(cell)
        set_column_width(ws, col_idx, col_name)
        headers[col_name] = col_idx

    for row_num in range(2, ws.max_row + 1):
        style_body_row(ws, row_num)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for col_name in ("Цена", "Цена со скидкой"):
        if col_name in headers:
            col_idx = headers[col_name]
            for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
                for c in col_cells:
                    c.number_format = '#,##0'

    if "Общая площадь" in headers:
        col_idx = headers["Общая площадь"]
        for col_cells in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in col_cells:
                c.number_format = '0.00'

    return headers


def update_workbook(source_xml: str, excel_path: str, output_path: str) -> dict[str, object]:
    """Обновляет Excel через полную пересборку строк по актуальному XML.

    Это более надежная логика, чем физическое удаление отдельных строк в существующем
    листе. Мы читаем старую таблицу как базу ручных данных, затем создаем новый лист
    строго по актуальному списку ID из XML. Поэтому пустые строки, «дыры» и хвосты
    форматирования Excel не могут попасть в итоговую таблицу.
    """
    xml_bytes = load_xml(source_xml)
    feed_rows, feed_columns, feed_format = parse_universal(xml_bytes)

    feed_order: list[str] = []
    feed_by_id: dict[str, dict[str, object]] = {}
    for row in feed_rows:
        row_id = normalize_id(row.get("ID квартиры"))
        if not row_id:
            continue
        feed_order.append(row_id)
        feed_by_id[row_id] = row

    if not feed_by_id:
        raise ValueError("В XML не найдено ни одной квартиры с ID. Обновление остановлено.")

    old_wb = load_workbook(excel_path)
    old_ws = old_wb.active
    old_headers = header_map(old_ws)
    rows_without_id_in_old_excel = count_rows_without_id_in_source(old_ws, old_headers)
    old_data_by_id = build_existing_data_by_id(old_ws, old_headers)

    old_ids = set(old_data_by_id.keys())
    xml_ids = set(feed_by_id.keys())
    ids_to_delete = old_ids - xml_ids
    ids_to_add = xml_ids - old_ids

    output_columns = build_output_columns(old_headers, feed_columns)

    new_wb = type(old_wb)()
    new_ws = new_wb.active
    new_ws.title = old_ws.title or "Лоты"

    # Заголовки.
    for col_idx, col_name in enumerate(output_columns, start=1):
        new_ws.cell(row=1, column=col_idx).value = col_name

    changed_rows = 0
    changed_cells = 0

    for out_row_num, row_id in enumerate(feed_order, start=2):
        feed_row = feed_by_id[row_id]
        old_row = old_data_by_id.get(row_id)
        output_values = make_output_row(row_id, feed_row, old_row, output_columns)

        # Считаем изменения только для существующих квартир.
        if old_row:
            row_changed = False
            for col_name, new_value in zip(output_columns, output_values):
                old_value = old_row.get(col_name, "")
                # Сравниваем только те поля, которые реально обновляются из XML.
                if col_name == "ID квартиры" or col_name in DYNAMIC_UPDATE_FIELDS or col_name.startswith("Изображение "):
                    if normalize_compare_value(old_value) != normalize_compare_value(new_value):
                        changed_cells += 1
                        row_changed = True
            if row_changed:
                changed_rows += 1

        for col_idx, value in enumerate(output_values, start=1):
            new_ws.cell(row=out_row_num, column=col_idx).value = value

    new_headers = apply_formatting_to_new_workbook(new_ws, output_columns)
    excel_count_after = count_rows_with_id(new_ws, new_headers)

    # Если исходная книга содержала другие листы, переносим их как есть после основного листа.
    # Это задел на будущее: вдруг пользователь добавит справочный лист.
    for sheet_name in old_wb.sheetnames[1:]:
        source_ws = old_wb[sheet_name]
        target_ws = new_wb.create_sheet(title=sheet_name)
        for row in source_ws.iter_rows():
            for cell in row:
                target = target_ws[cell.coordinate]
                target.value = cell.value
                if cell.has_style:
                    target._style = copy.copy(cell._style)
                if cell.number_format:
                    target.number_format = cell.number_format
                if cell.alignment:
                    target.alignment = copy.copy(cell.alignment)
                if cell.border:
                    target.border = copy.copy(cell.border)
                if cell.fill:
                    target.fill = copy.copy(cell.fill)
                if cell.font:
                    target.font = copy.copy(cell.font)

    new_wb.save(output_path)

    return {
        "output": output_path,
        "feed_format": feed_format,
        "xml_count": len(feed_by_id),
        "excel_count_before": len(old_ids),
        "deleted": len(ids_to_delete),
        "deleted_empty_rows": 0,
        "deleted_rows_without_id": rows_without_id_in_old_excel,
        "added": len(ids_to_add),
        "updated_rows": changed_rows,
        "updated_cells": changed_cells,
        "excel_count_after": excel_count_after,
        "rebuild_mode": True,
    }

def main() -> int:
    parser = argparse.ArgumentParser(
        description="Универсальное обновление Excel-таблицы по XML-фиду Яндекс.Недвижимость или Домклик/Kvartirogramma"
    )
    parser.add_argument("--source", required=True, help="Путь к XML-файлу или URL XML-фида")
    parser.add_argument("--excel", required=True, help="Путь к существующему XLSX, который нужно обновить")
    parser.add_argument("--output", required=True, help="Путь для сохранения обновленного XLSX")
    args = parser.parse_args()

    try:
        result = update_workbook(args.source, args.excel, args.output)
    except Exception as exc:
        print(f"Ошибка: {exc}", file=sys.stderr)
        return 1

    print(f"Версия скрипта обновления: {SCRIPT_VERSION}")
    print(f"Готово: {result['output']}")
    print(f"Формат фида: {result['feed_format']}")
    print(f"Квартир в XML: {result['xml_count']}")
    print(f"Квартир в Excel до обновления: {result['excel_count_before']}")
    print(f"Удалено строк: {result['deleted']}")
    print(f"Удалено полностью пустых строк: {result['deleted_empty_rows']}")
    print(f"Удалено строк без ID: {result['deleted_rows_without_id']}")
    print(f"Добавлено строк: {result['added']}")
    print(f"Обновлено квартир: {result['updated_rows']}")
    print(f"Обновлено ячеек: {result['updated_cells']}")
    print(f"Квартир в Excel после обновления: {result['excel_count_after']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
