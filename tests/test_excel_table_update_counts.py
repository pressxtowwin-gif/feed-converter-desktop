from pathlib import Path

import pytest

openpyxl = pytest.importorskip("openpyxl")
from openpyxl import load_workbook

from core.excel_table import create_excel_from_feed_rows, update_excel_from_feed_rows


def _feed_rows(count: int = 5) -> list[dict[str, str]]:
    return [
        {
            "ID квартиры": str(idx),
            "Номер квартиры": f"{100 + idx}",
            "Цена": str(idx * 1000),
            "Описание": f"description {idx}",
            "Планировка": f"plan-{idx}",
            "Пользовательское": f"manual {idx}",
        }
        for idx in range(1, count + 1)
    ]


def _columns() -> list[str]:
    return ["ID квартиры", "Номер квартиры", "Цена", "Описание", "Планировка", "Пользовательское"]


def test_update_counts_restored_feed_controlled_cells(tmp_path: Path) -> None:
    columns = ["ID квартиры", "Номер квартиры", "Цена", "Описание", "Планировка", "Пользовательское"]
    original_rows = [
        {
            "ID квартиры": "1",
            "Номер квартиры": "101",
            "Цена": "1000",
            "Описание": "old description",
            "Планировка": "plan-a",
            "Пользовательское": "manual note",
        },
        {
            "ID квартиры": "2",
            "Номер квартиры": "102",
            "Цена": "2000",
            "Описание": "description 2",
            "Планировка": "plan-b",
            "Пользовательское": "manual note 2",
        },
    ]
    excel_path = tmp_path / "lots.xlsx"
    create_excel_from_feed_rows(original_rows, columns, excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    # Existing row 1: clear/change dynamic feed-controlled cells.
    ws["C2"] = None
    ws["D2"] = "user changed dynamic description"
    # Existing row 2: clear a manual/static field. It is restored only because it is blank.
    ws["B3"] = None
    # Manual custom field must stay untouched and not be counted.
    ws["F3"] = "edited manual note"
    wb.save(excel_path)

    feed_rows = [
        {
            "ID квартиры": "1",
            "Номер квартиры": "101",
            "Цена": "1000",
            "Описание": "old description",
            "Планировка": "plan-a",
            "Пользовательское": "manual note",
        },
        {
            "ID квартиры": "2",
            "Номер квартиры": "102",
            "Цена": "2000",
            "Описание": "description 2",
            "Планировка": "plan-b",
            "Пользовательское": "manual note 2",
        },
    ]
    output_path = tmp_path / "updated.xlsx"

    result = update_excel_from_feed_rows(feed_rows, columns, excel_path, output_path)

    assert result["repaired_excel_cells"] == 3
    assert result["repaired_excel_apartments"] == 2
    assert result["updated_cells"] == 3
    assert result["updated_apartments"] == 2
    assert result["added"] == 0
    assert result["deleted"] == 0
    assert result["excel_rows_before"] == 2
    assert result["excel_rows_after"] == 2

    updated = load_workbook(output_path).active
    assert updated["C2"].value == "1000"
    assert updated["D2"].value == "old description"
    assert updated["B3"].value == "102"
    assert updated["F3"].value == "edited manual note"


def test_update_does_not_count_unchanged_existing_rows(tmp_path: Path) -> None:
    columns = ["ID квартиры", "Цена", "Описание"]
    rows = [{"ID квартиры": "1", "Цена": "1000", "Описание": "same"}]
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(rows, columns, excel_path)

    result = update_excel_from_feed_rows(rows, columns, excel_path, output_path)

    assert result["updated_cells"] == 0
    assert result["updated_apartments"] == 0
    assert result["added"] == 0
    assert result["deleted"] == 0


def test_clearing_cells_in_existing_rows_counts_updated_apartments_not_added(tmp_path: Path) -> None:
    columns = _columns()
    rows = _feed_rows(5)
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(rows, columns, excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws["C2"] = None
    ws["D3"] = None
    ws["E4"] = "changed plan"
    wb.save(excel_path)

    result = update_excel_from_feed_rows(rows, columns, excel_path, output_path)

    assert result["repaired_excel_apartments"] == 3
    assert result["repaired_excel_cells"] == 3
    assert result["updated_apartments"] == 3
    assert result["updated_cells"] == 3
    assert result["added"] == 0
    assert result["deleted"] == 0
    assert result["feed_rows"] == 5
    assert result["excel_rows_before"] == 5
    assert result["excel_rows_after"] == 5


def test_deleting_full_rows_by_apartment_id_counts_as_restored_excel_rows(tmp_path: Path) -> None:
    columns = _columns()
    rows = _feed_rows(5)
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(rows, columns, excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws.delete_rows(3, 2)
    wb.save(excel_path)

    result = update_excel_from_feed_rows(rows, columns, excel_path, output_path)

    assert result["restored_excel_rows"] == 2
    assert result["restored_excel_ids"] == ["2", "3"]
    assert result["added"] == 0
    assert result["added_ids"] == []
    assert result["deleted"] == 0
    assert result["excel_rows_before"] == 3
    assert result["excel_rows_after"] == 5


def test_changing_price_repairs_excel_but_not_feed_price_changes(tmp_path: Path) -> None:
    columns = _columns()
    rows = _feed_rows(4)
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(rows, columns, excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws["C2"] = "999"
    ws["C4"] = "2999"
    wb.save(excel_path)

    result = update_excel_from_feed_rows(rows, columns, excel_path, output_path)

    assert result["prices_changed"] == 0
    assert len(result["price_changes"]) == 0
    assert result["repaired_excel_apartments"] == 2
    assert result["updated_apartments"] >= 2
    assert result["updated_cells"] >= 2
    assert result["added"] == 0


def test_unchanged_file_counts_no_updates_or_additions(tmp_path: Path) -> None:
    columns = _columns()
    rows = _feed_rows(4)
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(rows, columns, excel_path)

    result = update_excel_from_feed_rows(rows, columns, excel_path, output_path)

    assert result["updated_apartments"] == 0
    assert result["updated_cells"] == 0
    assert result["restored_excel_rows"] == 0
    assert result["repaired_excel_apartments"] == 0
    assert result["repaired_excel_cells"] == 0
    assert result["added"] == 0
    assert result["deleted"] == 0


def test_changed_feed_business_counters_use_previous_feed_state(tmp_path: Path) -> None:
    columns = _columns()
    previous_rows = _feed_rows(4)
    current_rows = _feed_rows(4)
    current_rows[1]["Цена"] = "2500"
    current_rows.append({
        "ID квартиры": "5",
        "Номер квартиры": "105",
        "Цена": "5000",
        "Описание": "description 5",
        "Планировка": "plan-5",
        "Пользовательское": "manual 5",
    })
    current_rows = [row for row in current_rows if row["ID квартиры"] != "3"]
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(previous_rows, columns, excel_path)

    result = update_excel_from_feed_rows(
        current_rows,
        columns,
        excel_path,
        output_path,
        previous_feed_rows=previous_rows,
    )

    assert result["added"] == 1
    assert result["added_ids"] == ["5"]
    assert result["deleted"] == 1
    assert result["deleted_ids"] == ["3"]
    assert result["prices_changed"] == 1
    assert result["restored_excel_rows"] == 0


def test_one_apartment_with_several_cleared_cells_has_separate_repair_aggregates(tmp_path: Path) -> None:
    columns = _columns()
    rows = _feed_rows(3)
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(rows, columns, excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws["C3"] = None
    ws["D3"] = None
    ws["E3"] = None
    wb.save(excel_path)

    result = update_excel_from_feed_rows(rows, columns, excel_path, output_path)

    assert result["repaired_excel_apartments"] == 1
    assert result["repaired_excel_cells"] == 3
    assert result["repaired_excel_apartment_ids"] == ["2"]
    assert result["repaired_excel_cells_by_apartment"] == {"2": 3}
    assert result["restored_excel_rows"] == 0
    assert result["restored_excel_ids"] == []
    assert result["added"] == 0
    assert result["added_ids"] == []


def test_feed_changes_and_excel_damage_are_accounted_separately(tmp_path: Path) -> None:
    columns = _columns()
    previous_rows = _feed_rows(4)
    current_rows = _feed_rows(4)
    current_rows[1]["Цена"] = "2500"
    current_rows[2]["Описание"] = "new feed description"
    current_rows.append({
        "ID квартиры": "5",
        "Номер квартиры": "105",
        "Цена": "5000",
        "Описание": "description 5",
        "Планировка": "plan-5",
        "Пользовательское": "manual 5",
    })
    current_rows = [row for row in current_rows if row["ID квартиры"] != "4"]
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(previous_rows, columns, excel_path)

    wb = load_workbook(excel_path)
    ws = wb.active
    ws["C2"] = "999"  # Excel-only price damage for apartment 1.
    ws["D3"] = None  # Damaged same apartment that also has a real feed price change.
    wb.save(excel_path)

    result = update_excel_from_feed_rows(
        current_rows,
        columns,
        excel_path,
        output_path,
        previous_feed_rows=previous_rows,
    )

    assert result["added"] == 1
    assert result["added_ids"] == ["5"]
    assert result["deleted"] == 1
    assert result["deleted_ids"] == ["4"]
    assert result["prices_changed"] == 1
    assert result["price_changes"] == [{"ID квартиры": "2", "old_price": 2000, "new_price": 2500}]
    assert result["repaired_excel_apartments"] == 3
    assert result["repaired_excel_cells"] == 3
    assert result["repaired_excel_apartment_ids"] == ["1", "2", "3"]
    assert result["repaired_excel_cells_by_apartment"] == {"1": 1, "2": 1, "3": 1}
    assert result["restored_excel_rows"] == 0
    assert result["restored_excel_ids"] == []

    updated = load_workbook(output_path).active
    assert updated["C2"].value == 1000
    assert updated["C3"].value == 2500
    assert updated["D4"].value == "new feed description"


def test_no_previous_feed_archive_treats_current_feed_as_baseline(tmp_path: Path) -> None:
    columns = _columns()
    rows = _feed_rows(2)
    excel_path = tmp_path / "lots.xlsx"
    output_path = tmp_path / "updated.xlsx"
    create_excel_from_feed_rows(rows, columns, excel_path)

    result = update_excel_from_feed_rows(rows, columns, excel_path, output_path, previous_feed_rows=None)

    assert result["added"] == 0
    assert result["deleted"] == 0
    assert result["prices_changed"] == 0
    assert result["restored_excel_rows"] == 0


def _yandex_xml(*ids: str) -> bytes:
    offers = "".join(
        f"""
        <offer internal-id=\"{apt_id}\">
            <building-name>ЖК Test</building-name>
            <apartments>{apt_id}</apartments>
            <price><value>{int(apt_id) * 1000}</value></price>
        </offer>
        """
        for apt_id in ids
    )
    return f"<realty-feed>{offers}</realty-feed>".encode()


def test_previous_feed_rows_returns_none_when_no_archive_exists(tmp_path: Path) -> None:
    from core.project_service import _previous_feed_rows

    project_dir = tmp_path / "project"
    (project_dir / "xml" / "archive").mkdir(parents=True)

    assert _previous_feed_rows(project_dir) is None


def test_previous_feed_rows_uses_latest_archived_current_xml(tmp_path: Path) -> None:
    from core.project_service import _previous_feed_rows

    project_dir = tmp_path / "project"
    archive_dir = project_dir / "xml" / "archive"
    archive_dir.mkdir(parents=True)
    older = archive_dir / "current_20260101000000.xml"
    newer = archive_dir / "current_20260201000000.xml"
    ignored = archive_dir / "downloaded_20260301000000.xml"
    older.write_bytes(_yandex_xml("1"))
    newer.write_bytes(_yandex_xml("2", "3"))
    ignored.write_bytes(_yandex_xml("9"))

    rows = _previous_feed_rows(project_dir)

    assert [row["ID квартиры"] for row in rows or []] == ["2", "3"]
